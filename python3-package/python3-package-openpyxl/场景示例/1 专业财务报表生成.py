from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import datetime


def create_financial_report():
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "财务报表 Q3"

    # ===== 1. 报表标题和元信息 =====
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "2023年第三季度财务报表"
    title_cell.font = Font(name='微软雅黑', size=18, bold=True, color="003366")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 报表日期
    ws['A2'] = "生成日期:"
    ws['B2'] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws['E2'] = "单位: 人民币(万元)"
    ws['E2'].font = Font(italic=True)

    # ===== 2. 表头设置 =====
    headers = ["项目", "7月", "8月", "9月", "季度总计"]

    # 设置表头样式
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_border = Border(
        bottom=Side(border_style="medium", color="000000"),
        top=Side(border_style="thin", color="000000")
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')

    # ===== 3. 数据填充 =====
    financial_data = [
        ["营业收入", 1520, 1680, 1850],
        ["营业成本", 850, 920, 980],
        ["毛利润", 670, 760, 870],
        ["销售费用", 120, 135, 150],
        ["管理费用", 80, 85, 90],
        ["研发费用", 95, 100, 110],
        ["营业利润", 375, 440, 520],
        ["所得税", 56, 66, 78],
        ["净利润", 319, 374, 442]
    ]

    # 数据样式
    data_font = Font(name='Arial', size=11)
    number_format = '#,##0_);[Red](#,##0)'

    for row_idx, row_data in enumerate(financial_data, 5):
        # 写入项目名称
        ws.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)

        # 写入月度数据
        for col_idx in range(2, 5):
            ws.cell(row=row_idx, column=col_idx, value=row_data[col_idx - 1])
            ws.cell(row=row_idx, column=col_idx).number_format = number_format
            ws.cell(row=row_idx, column=col_idx).font = data_font

        # 计算季度总计
        total_formula = f"=SUM(B{row_idx}:D{row_idx})"
        ws.cell(row=row_idx, column=5, value=total_formula)
        ws.cell(row=row_idx, column=5).number_format = number_format
        ws.cell(row=row_idx, column=5).font = Font(bold=True, color="0000FF")

    # ===== 4. 公式和计算行 =====
    # 添加增长行
    ws.append(["增长率(%)", "", "", "", ""])
    for col in range(2, 6):
        current_row = len(financial_data) + 5
        prev_row = current_row - len(financial_data)
        if col < 5:
            formula = f"=IF(B{prev_row}<>0, (B{current_row - 1}-B{prev_row})/ABS(B{prev_row}), 0)"
            ws.cell(row=current_row, column=col, value=formula)
        else:
            formula = f"=IF(E{prev_row}<>0, (E{current_row - 1}-E{prev_row})/ABS(E{prev_row}), 0)"
            ws.cell(row=current_row, column=col, value=formula)

        ws.cell(row=current_row, column=col).number_format = "0.00%"
        ws.cell(row=current_row, column=col).font = Font(color="009900" if col == 5 else "000000")

    # ===== 5. 格式美化 =====
    # 设置列宽
    for col in range(1, 6):
        if col == 1:
            ws.column_dimensions[get_column_letter(col)].width = 20
        else:
            ws.column_dimensions[get_column_letter(col)].width = 15

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=4, max_row=len(financial_data) + 5, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    # ===== 6. 添加公司Logo =====
    try:
        logo = Image("company_logo.png")  # 替换为实际Logo路径
        logo.width = 100
        logo.height = 50
        ws.add_image(logo, "F1")
    except:
        print("Logo未找到，跳过添加")

    # ===== 7. 保存文件 =====
    filename = f"财务报表_Q3_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"财务报表已生成: {filename}")

if __name__ == '__main__':
    # 运行报表生成
    create_financial_report()