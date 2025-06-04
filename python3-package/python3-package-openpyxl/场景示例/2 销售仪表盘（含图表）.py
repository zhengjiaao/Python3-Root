from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_sales_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "销售仪表盘"

    # ===== 1. 准备数据 =====
    products = ["产品A", "产品B", "产品C", "产品D"]
    regions = ["华东", "华南", "华北", "西部"]

    # 生成随机销售数据
    import random
    data = []
    for product in products:
        row = [product]
        for region in regions:
            row.append(random.randint(50, 200))
        row.append(sum(row[1:]))  # 总计
        data.append(row)

    # ===== 2. 写入数据 =====
    # 表头
    headers = ["产品"] + regions + ["总计"]
    ws.append(headers)

    # 数据行
    for row in data:
        ws.append(row)

    # ===== 3. 创建图表 =====
    # 创建柱状图
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "区域产品销售对比"
    chart1.y_axis.title = "销售量"
    chart1.x_axis.title = "产品"

    # 设置数据范围
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=len(regions) + 1, max_row=len(products) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(products) + 1)

    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)

    # 将图表添加到工作表
    ws.add_chart(chart1, "F2")

    # 创建总计柱状图
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.title = "产品总销量"

    total_ref = Reference(ws, min_col=len(headers), min_row=2, max_row=len(products) + 1)
    chart2.add_data(total_ref)
    chart2.set_categories(cats_ref)

    ws.add_chart(chart2, "F20")

    # ===== 4. 格式设置 =====
    # 表头样式
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 数据行格式
    for row in range(2, len(products) + 2):
        # 设置产品名格式
        ws.cell(row=row, column=1).font = Font(bold=True)

        # 设置总计列格式
        total_cell = ws.cell(row=row, column=len(headers))
        total_cell.font = Font(bold=True, color="0070C0")
        total_cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")

    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # ===== 5. 添加KPI指标 =====
    kpi_row = len(products) + 3

    # 计算总销售量
    ws.cell(row=kpi_row, column=1, value="总销售量:").font = Font(bold=True)
    ws.cell(row=kpi_row, column=2, value=f"=SUM(E2:E{len(products) + 1})").font = Font(size=14, bold=True)

    # 计算最畅销产品
    ws.cell(row=kpi_row + 1, column=1, value="最畅销产品:").font = Font(bold=True)
    ws.cell(row=kpi_row + 1, column=2,
            value=f"=INDEX(A2:A{len(products) + 1},MATCH(MAX(F2:F{len(products) + 1}),F2:F{len(products) + 1},0))")
    ws.cell(row=kpi_row + 1, column=2).font = Font(color="00B050", bold=True)

    # ===== 6. 保存文件 =====
    wb.save("certificate_template.xlsx")
    print("销售仪表盘已生成")

if __name__ == '__main__':
    # 创建仪表盘
    create_sales_dashboard()