from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar
import random


def generate_attendance_report(month, year):
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月考勤"

    # ===== 1. 创建日历标题 =====
    # 获取月份天数
    _, num_days = calendar.monthrange(year, month)

    # 创建表头
    headers = ["员工ID", "姓名", "部门"] + [str(day) for day in range(1, num_days + 1)] + ["出勤率"]
    ws.append(headers)

    # ===== 2. 添加员工数据 =====
    # 示例员工数据
    employees = [
        ["E001", "张三", "技术部"],
        ["E002", "李四", "市场部"],
        ["E003", "王五", "财务部"],
        ["E004", "赵六", "技术部"],
        ["E005", "钱七", "人事部"]
    ]

    for emp in employees:
        # 添加基础信息
        ws.append(emp + [""] * (num_days + 1))

    # ===== 3. 设置考勤格式 =====
    # 定义状态样式
    status_styles = {
        "✓": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 出勤
        "△": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 迟到
        "✗": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 缺勤
        "○": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 休假
    }

    # 填充考勤数据
    for row in range(2, len(employees) + 2):
        # 添加出勤率公式
        attendance_col = len(headers)
        attendance_formula = f'=COUNTIF(D{row}:{get_column_letter(len(headers) - 1)}{row},"✓")/{num_days}'
        ws.cell(row=row, column=attendance_col, value=attendance_formula)
        ws.cell(row=row, column=attendance_col).number_format = "0.00%"

        # 填充每日考勤状态
        for col in range(4, 4 + num_days):
            # 随机生成考勤状态
            status = random.choice(["✓", "✓", "✓", "✓", "△", "✗", "○"])
            cell = ws.cell(row=row, column=col, value=status)
            cell.fill = status_styles.get(status, PatternFill())
            cell.alignment = Alignment(horizontal='center')

    # ===== 4. 设置周末格式 =====
    # 设置周末背景色
    weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col in range(4, 4 + num_days):
        col_date = datetime(year, month, col - 3)
        if col_date.weekday() >= 5:  # 5=周六, 6=周日
            for row in range(2, len(employees) + 2):
                ws.cell(row=row, column=col).fill = weekend_fill

    # ===== 5. 设置表头格式 =====
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 冻结窗格（固定前三列）
    ws.freeze_panes = "D2"

    # ===== 6. 自动调整列宽 =====
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        # 标题宽度
        if col_idx == 1:
            max_length = 8
        elif col_idx == 2:
            max_length = 10
        elif col_idx == 3:
            max_length = 12
        elif col_idx > 3 and col_idx <= 3 + num_days:
            max_length = 4
        else:
            max_length = 10

        ws.column_dimensions[col_letter].width = max_length

    # ===== 7. 添加统计表 =====
    ws2 = wb.create_sheet("部门统计")
    ws2.append(["部门", "员工数", "平均出勤率", "缺勤次数"])

    # 部门统计数据
    departments = set(emp[2] for emp in employees)
    for dept in departments:
        # 统计部门数据
        ws2.append([
            dept,
            f'=COUNTIF(\'{ws.title}\'!C:C, "{dept}")',
            f'=AVERAGEIF(\'{ws.title}\'!C:C, "{dept}", \'{ws.title}\'!{get_column_letter(len(headers))}:{get_column_letter(len(headers))})',
            f'=SUMIF(\'{ws.title}\'!C:C, "{dept}", \'{ws.title}\'!{get_column_letter(len(headers) - 1)}:{get_column_letter(len(headers) - 1)})'
        ])

    # 设置统计表格式
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # ===== 8. 保存文件 =====
    filename = f"{year}年{month}月考勤表.xlsx"
    wb.save(filename)
    print(f"考勤表已生成: {filename}")


if __name__ == '__main__':
    # 生成2023年10月考勤表
    generate_attendance_report(month=10, year=2023)