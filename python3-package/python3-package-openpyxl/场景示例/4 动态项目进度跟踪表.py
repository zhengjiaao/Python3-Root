from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
import random


def create_project_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "项目进度"

    # ===== 1. 项目表头 =====
    headers = [
        "项目ID", "项目名称", "负责人", "开始日期", "截止日期",
        "当前进度", "状态", "延迟天数", "风险等级"
    ]
    ws.append(headers)

    # ===== 2. 示例项目数据 =====
    projects = [
        ["P1001", "客户管理系统升级", "张三", "2023-10-01", "2023-11-15"],
        ["P1002", "移动APP开发", "李四", "2023-09-15", "2023-12-20"],
        ["P1003", "数据平台迁移", "王五", "2023-10-10", "2023-11-30"],
        ["P1004", "网络安全加固", "赵六", "2023-08-20", "2023-10-31"],
        ["P1005", "AI客服系统", "钱七", "2023-11-01", "2024-01-15"]
    ]

    # ===== 3. 填充项目数据 =====
    from datetime import datetime, timedelta

    for project in projects:
        # 随机生成进度和状态
        progress = random.randint(20, 95)

        # 计算状态和延迟
        end_date = datetime.strptime(project[4], "%Y-%m-%d")
        today = datetime.now()

        if progress >= 100:
            status = "已完成"
            delay = 0
        elif end_date < today:
            status = "已延期"
            delay = (today - end_date).days
        else:
            status = "进行中"
            delay = 0

        # 确定风险等级
        if delay > 7:
            risk = "高"
        elif delay > 3:
            risk = "中"
        elif progress < 30 and (end_date - today).days < 14:
            risk = "中"
        else:
            risk = "低"

        # 添加完整行
        ws.append(project + [progress / 100, status, delay, risk])

    # ===== 4. 设置进度条格式 =====
    # 添加数据条规则
    progress_rule = DataBarRule(
        start_type='num', start_value=0,
        end_type='num', end_value=1,
        color="63C384"  # 绿色
    )
    ws.conditional_formatting.add(
        f"F2:F{len(projects) + 1}",
        progress_rule
    )

    # 设置进度百分比格式
    for row in range(2, len(projects) + 2):
        ws.cell(row=row, column=6).number_format = "0%"

    # ===== 5. 设置状态颜色 =====
    status_colors = {
        "已完成": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "进行中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "已延期": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    }

    for row in range(2, len(projects) + 2):
        status = ws.cell(row=row, column=7).value
        if status in status_colors:
            ws.cell(row=row, column=7).fill = status_colors[status]

    # ===== 6. 设置风险等级格式 =====
    risk_colors = {
        "高": Font(color="FF0000", bold=True),
        "中": Font(color="FF9900", bold=True),
        "低": Font(color="00B050")
    }

    for row in range(2, len(projects) + 2):
        risk = ws.cell(row=row, column=9).value
        if risk in risk_colors:
            ws.cell(row=row, column=9).font = risk_colors[risk]

    # ===== 7. 表头格式 =====
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # ===== 8. 自动调整列宽 =====
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # 设置最小宽度
        if col_idx in [2, 3]:
            max_length = max(max_length, 15)

        ws.column_dimensions[col_letter].width = max_length + 2

    # ===== 9. 添加筛选器 =====
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(projects) + 1}"

    # ===== 10. 添加项目统计 =====
    stats_row = len(projects) + 3
    ws.cell(row=stats_row, column=1, value="项目统计").font = Font(bold=True, size=14)

    # 统计公式
    stats_labels = ["总项目数", "进行中", "已完成", "已延期", "高风险项目"]
    stats_formulas = [
        f"=COUNTA(A2:A{len(projects) + 1})",
        f'=COUNTIF(G2:G{len(projects) + 1}, "进行中")',
        f'=COUNTIF(G2:G{len(projects) + 1}, "已完成")',
        f'=COUNTIF(G2:G{len(projects) + 1}, "已延期")',
        f'=COUNTIF(I2:I{len(projects) + 1}, "高")'
    ]

    for i, (label, formula) in enumerate(zip(stats_labels, stats_formulas)):
        ws.cell(row=stats_row + 1, column=i * 2 + 1, value=label).font = Font(bold=True)
        cell = ws.cell(row=stats_row + 1, column=i * 2 + 2, value=formula)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")

    # ===== 11. 保存文件 =====
    wb.save("项目进度跟踪表.xlsx")
    print("项目进度跟踪表已生成")

if __name__ == '__main__':
    # 创建项目跟踪表
    create_project_tracker()