from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os


def generate_certificates():
    # 加载模板文件
    template_path = "template/certificate_template.xlsx"
    if not os.path.exists(template_path):
        print(f"模板文件不存在: {template_path}")
        return

    # 员工数据
    employees = [
        {"name": "张三", "department": "技术部", "award": "年度最佳员工"},
        {"name": "李四", "department": "市场部", "award": "销售冠军"},
        {"name": "王五", "department": "设计部", "award": "创新设计奖"},
        {"name": "赵六", "department": "客服部", "award": "客户满意奖"},
        {"name": "钱七", "department": "财务部", "award": "成本控制奖"}
    ]

    # 创建输出目录
    output_dir = "tmp/certificates"
    os.makedirs(output_dir, exist_ok=True)

    # 为每位员工生成证书
    for emp in employees:
        # 加载模板
        wb = load_workbook(template_path)
        ws = wb.active

        # 填充数据
        ws["C6"].value = emp["name"]  # 姓名
        ws["C8"].value = emp["department"]  # 部门
        ws["C10"].value = emp["award"]  # 奖项

        # 设置日期
        from datetime import datetime
        ws["G12"].value = datetime.now().strftime("%Y年%m月%d日")

        # 设置姓名样式
        ws["C6"].font = Font(name="楷体", size=20, bold=True)
        ws["C6"].alignment = Alignment(horizontal='center')

        # 保存证书
        filename = os.path.join(output_dir, f"{emp['name']}_获奖证书.xlsx")
        wb.save(filename)
        print(f"已生成: {filename}")

    print(f"\n共生成 {len(employees)} 份证书")


if __name__ == '__main__':
    # 生成证书
    generate_certificates()