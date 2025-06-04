# openpyxl 示例：创建带样式的报表
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

if __name__ == '__main__':
    print("openpyxl 示例：创建带样式的报表")

    wb = Workbook()
    ws = wb.active

    # 写入标题（带样式）
    title_cell = ws['A1']
    title_cell.value = "销售报告"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # 合并单元格
    ws.merge_cells('A1:D1')

    # 写入数据
    data = [["产品", "季度", "销量", "增长率"],
            ["A", "Q1", 1500, "15%"],
            ["B", "Q1", 2400, "22%"]]
    for row in data:
        ws.append(row)

    # 保存
    wb.save("sales_report.xlsx")