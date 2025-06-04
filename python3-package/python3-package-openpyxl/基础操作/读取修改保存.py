from openpyxl import load_workbook
from openpyxl.workbook import Workbook

if __name__ == '__main__':
    print("openpyxl 示例：读取、修改和保存 Excel 文件")
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 设置表头
    ws.append(["ID", "名称", "类别", "数量"])

    # 添加示例数据
    data = [
        [1, "苹果", "水果", 10],
        [2, "香蕉", "水果", 15],
        [3, "牛奶", "饮品", 8],
        [4, "面包", "食品", 5],
        [5, "鸡蛋", "食品", 20]
    ]

    # 写入数据
    for row in data:
        ws.append(row)

    # 保存为 data.xlsx
    wb.save('data.xlsx')

    print("Excel 文件已生成：data.xlsx")

    print("正在读取数据...")
    # 读取文件
    wb = load_workbook('data.xlsx')
    ws = wb.active

    # 遍历数据
    for row in ws.iter_rows(values_only=True):
        print(row)  # 输出每行数据元组

    print("正在修改数据...")
    # 修改单元格
    ws['B2'] = "新值"
    ws.cell(row=3, column=4).value = 100

    # 保存修改
    wb.save('modified.xlsx')
    print("Excel 文件已修改：modified.xlsx")