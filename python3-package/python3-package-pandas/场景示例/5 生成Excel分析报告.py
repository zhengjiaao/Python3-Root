import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import pandas as pd
import os

# 设置样式函数
def set_title_style(cell):
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal='center')

def set_header_style(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

def auto_adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            if isinstance(cell, openpyxl.cell.MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_length * 1.2 + 2, 50)

# 场景：将分析结果输出为专业Excel报告
if __name__ == '__main__':
    # 模拟数据
    product_analysis = pd.DataFrame({
        '产品': ['A', 'B', 'C'],
        '总销量': [100, 200, 150],
        '平均单价': [10.5, 20.0, 15.0],
        '订单数量': [10, 20, 15]
    })

    pivot_table = pd.DataFrame({
        '客户': ['Alice', 'Bob', 'Charlie'],
        '产品A': [3, 0, 2],
        '产品B': [1, 5, 0],
        '产品C': [0, 2, 4]
    })

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "销售分析报告"

    # 1. 添加标题
    ws['A1'] = "2023年销售分析报告"
    set_title_style(ws['A1'])
    ws.merge_cells('A1:F1')

    # 2. 添加产品分析表
    ws['A3'] = "产品分析"
    set_header_style(ws['A3'])

    headers = ["产品", "总销量", "平均单价", "订单数量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        set_header_style(cell)

    for i, row in product_analysis.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=5 + i, column=1 + j).value = value

    # 自动调整列宽
    auto_adjust_column_width(ws)

    # 3. 插入图表
    chart_positions = {
        'daily_sales.png': 'A10',
        'product_sales_pie.png': 'H10',
        'customer_profit.png': 'H30'
    }

    for image_file, position in chart_positions.items():
        try:
            img = Image(image_file)
            img.width = 500 if image_file == 'daily_sales.png' else 300
            img.height = 300
            ws.add_image(img, position)
        except FileNotFoundError:
            print(f"⚠️ 图像 {image_file} 未找到，跳过插入")

    # 4. 添加数据透视表
    ws2 = wb.create_sheet(title="客户购买明细")
    df = pd.DataFrame(pivot_table).reset_index()
    for r in df.itertuples():
        for c, value in enumerate(r[1:], 1):
            ws2.cell(row=r.Index + 1, column=c).value = value
    auto_adjust_column_width(ws2)

    # 保存报告
    output_path = "销售分析报告.xlsx"
    wb.save(output_path)
    print(f"✅ Excel分析报告已生成: {os.path.abspath(output_path)}")
